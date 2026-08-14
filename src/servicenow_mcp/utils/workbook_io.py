"""Generic Excel read/write for the documents this work actually produces.

Measured need, not imagined: one real session hand-wrote SEVEN openpyxl
scripts to query one tracking workbook (sheet list → header peek → date
filter → regex search → read rows in full), and the sign-off document at the
end was an 8,000-character heredoc whose first thirty lines — borders, fills,
fonts, widths, wrap — are identical every single time. The per-task part is
only the DATA. So:

- reads return compact, truncated-and-said-so row projections
- writes take a spec of sheets/rows/images and apply the house style
  themselves; the caller never authors styling code again
- fills write values into a copy of an existing form (the template is an
  input and is never written to)

Bulk discipline: workbook and image bytes stay on disk. Responses carry
values, counts and paths — with every truncation named, never silent.
"""

import logging
import os
from datetime import date, datetime
from datetime import time as dt_time
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

# Read caps: enough to answer "what is in here", small enough to never flood
# context. Every cap that bites is reported (`truncated`, `cell_truncated`).
MAX_ROWS = 100
DEFAULT_ROWS = 20
MAX_CELL_CHARS = 200

# Embedded image width in px; tall shots keep their aspect ratio.
IMAGE_WIDTH_PX = 440

_HOUSE = {
    "header_fill": "1F4E79",
    "header_font": "FFFFFF",
    "border": "BFBFBF",
    "label_fill": "DDEBF7",
}


class OpenpyxlUnavailable(RuntimeError):
    """openpyxl is not importable in this interpreter."""


def require_openpyxl() -> None:
    try:
        import openpyxl  # type: ignore[import-untyped]  # noqa: F401
    except ImportError as exc:  # pragma: no cover - environment dependent
        raise OpenpyxlUnavailable(
            "openpyxl is not installed. Install the excel extra "
            "(`pip install 'mfa-servicenow-mcp[excel]'`) or `pip install openpyxl`."
        ) from exc


# ---------------------------------------------------------------------------
# Read
# ---------------------------------------------------------------------------


def _cell_text(value: Any) -> Any:
    """Context-safe cell value: dates ISO, long text truncated with a marker."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return (
            value.strftime("%Y-%m-%d %H:%M")
            if (value.hour or value.minute)
            else value.strftime("%Y-%m-%d")
        )
    if isinstance(value, (date, dt_time)):
        return value.isoformat()
    if isinstance(value, str) and len(value) > MAX_CELL_CHARS:
        return value[:MAX_CELL_CHARS] + f"…(+{len(value) - MAX_CELL_CHARS}자)"
    return value


def _open_sheet(path: str, sheet: Optional[str]) -> Tuple[Any, Any]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet:
        if sheet not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet}' not found. Sheets: {wb.sheetnames}")
        return wb, wb[sheet]
    return wb, wb[wb.sheetnames[0]]


def _col_indexes(columns: Optional[str]) -> Optional[List[int]]:
    """'A,C,F' or '1,3,6' → zero-based indexes. None = all columns."""
    if not columns:
        return None
    from openpyxl.utils import column_index_from_string

    out: List[int] = []
    for token in str(columns).split(","):
        token = token.strip()
        if not token:
            continue
        if token.isdigit():
            out.append(int(token) - 1)
        else:
            out.append(column_index_from_string(token.upper()) - 1)
    return out or None


def _project(row: Tuple[Any, ...], picks: Optional[List[int]]) -> List[Any]:
    if picks is None:
        return [_cell_text(v) for v in row]
    return [_cell_text(row[i]) if i < len(row) else "" for i in picks]


def list_sheets(path: str) -> Dict[str, Any]:
    require_openpyxl()
    wb, _ = _open_sheet(path, None)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({"name": name, "rows": ws.max_row, "cols": ws.max_column})
    wb.close()
    return {"success": True, "file": path, "sheets": sheets}


def read_rows(
    path: str,
    sheet: Optional[str] = None,
    *,
    min_row: int = 1,
    limit: int = DEFAULT_ROWS,
    columns: Optional[str] = None,
) -> Dict[str, Any]:
    require_openpyxl()
    wb, ws = _open_sheet(path, sheet)
    picks = _col_indexes(columns)
    capped = max(1, min(int(limit or DEFAULT_ROWS), MAX_ROWS))
    rows: List[List[Any]] = []
    numbers: List[int] = []
    for index, row in enumerate(
        ws.iter_rows(min_row=max(1, min_row), values_only=True), start=max(1, min_row)
    ):
        if all(v is None for v in row):
            continue
        rows.append(_project(row, picks))
        numbers.append(index)
        if len(rows) >= capped:
            break
    total = ws.max_row
    wb.close()
    result: Dict[str, Any] = {
        "success": True,
        "sheet": ws.title,
        "rows": rows,
        "row_numbers": numbers,
        "sheet_rows_total": total,
    }
    last = numbers[-1] if numbers else min_row
    if last < total:
        # Named, never silent: there is more below what was returned.
        result["truncated"] = f"stopped at row {last} of {total} — continue with min_row={last + 1}"
    return result


def find_rows(
    path: str,
    query: str,
    sheet: Optional[str] = None,
    *,
    column: Optional[str] = None,
    min_row: int = 1,
    limit: int = DEFAULT_ROWS,
    columns: Optional[str] = None,
) -> Dict[str, Any]:
    """Regex search over rows — the whole row's text, or one column of it."""
    import re

    require_openpyxl()
    pattern = re.compile(query, re.IGNORECASE)
    wb, ws = _open_sheet(path, sheet)
    picks = _col_indexes(columns)
    scope = _col_indexes(column)
    capped = max(1, min(int(limit or DEFAULT_ROWS), MAX_ROWS))
    rows: List[List[Any]] = []
    numbers: List[int] = []
    scanned = matches = 0
    for index, row in enumerate(
        ws.iter_rows(min_row=max(1, min_row), values_only=True), start=max(1, min_row)
    ):
        scanned = index
        if scope is not None:
            haystack = " ".join(str(row[i]) for i in scope if i < len(row) and row[i] is not None)
        else:
            haystack = " ".join(str(v) for v in row if v is not None)
        if not haystack or not pattern.search(haystack):
            continue
        matches += 1
        if len(rows) < capped:
            rows.append(_project(row, picks))
            numbers.append(index)
    total = ws.max_row
    wb.close()
    result: Dict[str, Any] = {
        "success": True,
        "sheet": ws.title,
        "matches": matches,
        "rows": rows,
        "row_numbers": numbers,
        "scanned_to_row": scanned,
        "sheet_rows_total": total,
    }
    if matches > len(rows):
        result["truncated"] = (
            f"{matches} matches, first {len(rows)} returned — raise limit or narrow the query"
        )
    return result


# ---------------------------------------------------------------------------
# Write
# ---------------------------------------------------------------------------


def _pillow_available() -> bool:
    try:
        import PIL  # type: ignore[import-untyped]  # noqa: F401

        return True
    except ImportError:
        return False


def _thumbnail(source: str, work_dir: str) -> Optional[str]:
    """A resized copy for embedding; the original file is never touched."""
    try:
        from PIL import Image as PILImage

        with PILImage.open(source) as img:
            if img.width <= IMAGE_WIDTH_PX:
                return source
            ratio = IMAGE_WIDTH_PX / float(img.width)
            resized = img.resize((IMAGE_WIDTH_PX, max(1, int(img.height * ratio))))
            base = os.path.splitext(os.path.basename(source))[0]
            dest = os.path.join(work_dir, f"{base}.thumb.png")
            resized.save(dest, format="PNG")
            return dest
    except Exception as exc:  # noqa: BLE001 - a bad image degrades to a link
        logger.debug("Could not thumbnail %s: %s", source, exc)
        return None


def _place_image(ws: Any, spec: Dict[str, Any], work_dir: str) -> Optional[str]:
    """Embed one image. Returns an error string instead of raising."""
    path = str(spec.get("path") or "")
    cell = str(spec.get("cell") or "")
    if not path or not cell:
        return "image needs {path, cell}"
    if not os.path.isfile(path):
        return f"image not found: {path}"
    if not _pillow_available():
        ws[cell] = f"{os.path.basename(path)} (Pillow 미설치 — 파일 참조)"
        return f"Pillow missing — {os.path.basename(path)} linked as text"
    from openpyxl.drawing.image import Image as XLImage

    thumb = _thumbnail(path, work_dir)
    if thumb is None:
        ws[cell] = f"{os.path.basename(path)} (이미지 열기 실패 — 파일 참조)"
        return f"unreadable image — {os.path.basename(path)} linked as text"
    image = XLImage(thumb)
    image.anchor = cell
    ws.add_image(image)
    row = int("".join(ch for ch in cell if ch.isdigit()) or 1)
    ws.row_dimensions[row].height = max(
        ws.row_dimensions[row].height or 0, float(image.height) * 0.75 + 6
    )
    return None


def _apply_house_style(ws: Any, header_row: Optional[int], widths: Optional[List[float]]) -> None:
    """The thirty lines every hand-written script repeated, applied once."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color=_HOUSE["border"])
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if widths:
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = float(width)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if cell.value is None:
                continue
            cell.border = border
            if header_row and cell.row == header_row:
                cell.fill = PatternFill("solid", fgColor=_HOUSE["header_fill"])
                cell.font = Font(color=_HOUSE["header_font"], bold=True, size=10)
                cell.alignment = center
            else:
                cell.font = Font(size=10)
                cell.alignment = wrap


def write_workbook(spec: Dict[str, Any], output_path: str) -> Dict[str, Any]:
    """Build a styled workbook from data alone.

    spec = {"sheets": [{"title", "rows": [[...]], "header_row": 1,
                        "widths": [..], "label_col": true?, "images": [{path, cell}]}]}
    """
    require_openpyxl()
    import openpyxl

    sheets = list(spec.get("sheets") or [])
    if not sheets:
        return {"success": False, "error": "spec.sheets is empty — nothing to write."}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    work_dir = os.path.dirname(os.path.abspath(output_path)) or "."
    os.makedirs(work_dir, exist_ok=True)

    written = []
    image_notes: List[str] = []
    embedded = 0
    for index, sheet_spec in enumerate(sheets, start=1):
        title = str(sheet_spec.get("title") or f"Sheet{index}")[:31]
        ws = wb.create_sheet(title=title)
        rows = list(sheet_spec.get("rows") or [])
        for row in rows:
            ws.append(list(row) if isinstance(row, (list, tuple)) else [row])
        header_row = sheet_spec.get("header_row")
        _apply_house_style(
            ws,
            int(header_row) if header_row else None,
            [float(w) for w in (sheet_spec.get("widths") or [])] or None,
        )
        if sheet_spec.get("label_col"):
            from openpyxl.styles import PatternFill

            start = (int(header_row) + 1) if header_row else 1
            for row_index in range(start, ws.max_row + 1):
                if ws.cell(row_index, 1).value is not None:
                    ws.cell(row_index, 1).fill = PatternFill("solid", fgColor=_HOUSE["label_fill"])
        for image_spec in sheet_spec.get("images") or []:
            note = _place_image(ws, dict(image_spec), work_dir)
            if note:
                image_notes.append(f"{title}: {note}")
            else:
                embedded += 1
        written.append({"sheet": title, "rows": ws.max_row})

    wb.save(output_path)
    result: Dict[str, Any] = {
        "success": True,
        "file": os.path.abspath(output_path),
        "sheets": written,
        "embedded_images": embedded,
    }
    if image_notes:
        result["image_notes"] = image_notes
    return result


def fill_workbook(
    template_path: str,
    output_path: str,
    *,
    sheet: Optional[str] = None,
    cells: Optional[Dict[str, Any]] = None,
    rows_at: Optional[Dict[str, Any]] = None,
    images: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    """Fill a COPY of an existing form. The template file itself is an input.

    cells   = {"B3": "value", ...}
    rows_at = {"start_row": 5, "rows": [[...], ...], "start_col": 1}
    """
    require_openpyxl()
    import openpyxl

    if os.path.abspath(template_path) == os.path.abspath(output_path):
        return {
            "success": False,
            "error": "output_file must differ from path — the form is an input, not the deliverable.",
        }
    if not os.path.isfile(template_path):
        return {"success": False, "error": f"Template not found: {template_path}"}

    wb = openpyxl.load_workbook(template_path)
    if sheet and sheet not in wb.sheetnames:
        return {"success": False, "error": f"Sheet '{sheet}' not found. Sheets: {wb.sheetnames}"}
    ws = wb[sheet] if sheet else wb.active

    cells_written = 0
    for ref, value in (cells or {}).items():
        ws[str(ref)] = value
        cells_written += 1

    rows_written = 0
    if rows_at:
        start_row = int(rows_at.get("start_row") or (ws.max_row + 1))
        start_col = int(rows_at.get("start_col") or 1)
        for offset, row in enumerate(list(rows_at.get("rows") or [])):
            values = list(row) if isinstance(row, (list, tuple)) else [row]
            for col_offset, value in enumerate(values):
                ws.cell(row=start_row + offset, column=start_col + col_offset, value=value)
            rows_written += 1

    work_dir = os.path.dirname(os.path.abspath(output_path)) or "."
    os.makedirs(work_dir, exist_ok=True)
    image_notes: List[str] = []
    embedded = 0
    for image_spec in images or []:
        note = _place_image(ws, dict(image_spec), work_dir)
        if note:
            image_notes.append(note)
        else:
            embedded += 1

    wb.save(output_path)
    result: Dict[str, Any] = {
        "success": True,
        "file": os.path.abspath(output_path),
        "sheet": ws.title,
        "cells_written": cells_written,
        "rows_written": rows_written,
        "embedded_images": embedded,
    }
    if image_notes:
        result["image_notes"] = image_notes
    return result


__all__ = [
    "OpenpyxlUnavailable",
    "require_openpyxl",
    "list_sheets",
    "read_rows",
    "find_rows",
    "write_workbook",
    "fill_workbook",
]
